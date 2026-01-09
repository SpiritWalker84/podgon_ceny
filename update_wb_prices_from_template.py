#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Объединенный скрипт для обновления цен на Wildberries:
1. Скачивает актуальный Excel шаблон с рекомендуемыми ценами
2. Корректирует цены (устанавливает колонку J = N - 1)
3. Загружает скорректированные цены на WB через API
"""

import os
import sys
import time
from pathlib import Path
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv

# Загружаем переменные окружения
load_dotenv()
load_dotenv('.env')

# Добавляем текущую директорию в путь для импорта
sys.path.insert(0, str(Path(__file__).parent))

# Импортируем функции из других модулей
try:
    # Временно подавляем выполнение test_download_excel.py при импорте
    import test_download_excel
    download_excel_only = test_download_excel.download_excel_only
except ImportError:
    print("[ERROR] Не удалось импортировать download_excel_only из test_download_excel.py")
    sys.exit(1)

try:
    from update_prices import adjust_prices, find_wb_template_files
except ImportError:
    print("[ERROR] Не удалось импортировать adjust_prices из update_prices.py")
    sys.exit(1)

try:
    import requests
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] Необходимые библиотеки не установлены")
    print("Установите: py -m pip install requests openpyxl")
    sys.exit(1)


class Config:
    """Класс для хранения конфигурации из переменных окружения"""
    
    # API настройки
    WB_API_TOKEN: str = os.getenv('WB_API_TOKEN', '')
    
    # URL API
    PRICES_API_URL: str = "https://discounts-prices-api.wildberries.ru/api/v2"
    
    # Директория для работы
    TARGET_DIR: Path = Path(os.getenv('TARGET_DIR', str(Path.cwd())))
    
    @classmethod
    def validate(cls) -> None:
        """Проверяет, что все необходимые переменные окружения установлены"""
        if not cls.WB_API_TOKEN:
            raise ValueError("WB_API_TOKEN не установлен в .env файле")


def get_headers() -> Dict[str, str]:
    """Получить заголовки для API запросов"""
    token = Config.WB_API_TOKEN
    if not token:
        raise ValueError("WB_API_TOKEN не установлен в .env файле")
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }


def read_prices_from_excel_template(template_file: str) -> Dict[int, int]:
    """
    Читает цены из Excel шаблона WB.
    Читает nmID из колонки C и цену из колонки J.
    
    Args:
        template_file: Путь к Excel файлу шаблона
        
    Returns:
        Dict[int, int]: Словарь {nmID: price_in_rubles} из колонки J (цена в рублях, int)
    """
    prices = {}
    
    if not os.path.exists(template_file):
        print(f"[ERROR] Файл не найден: {template_file}")
        return prices
    
    print(f"[INFO] Читаю цены из файла: {os.path.basename(template_file)}")
    
    try:
        wb = load_workbook(template_file, data_only=True)
        ws = wb.active
        
        # Колонки в шаблоне WB:
        # Колонка C (3) - nmID
        # Колонка J (10) - цена (скорректированная)
        
        nmid_col = 3  # Колонка C
        price_col = 10  # Колонка J
        
        # Читаем данные начиная со второй строки (первая - заголовок)
        read_count = 0
        for row in range(2, ws.max_row + 1):
            try:
                nmid_cell = ws.cell(row=row, column=nmid_col)
                price_cell = ws.cell(row=row, column=price_col)
                
                # Пропускаем пустые строки
                if not nmid_cell.value or not price_cell.value:
                    continue
                
                try:
                    nmid = int(float(str(nmid_cell.value)))
                    price = float(str(price_cell.value))
                    
                    # WB API работает с ценами в рублях (int), не в копейках
                    # Округляем до целого числа рублей
                    price_rubles = int(round(price))
                    
                    if nmid > 0 and price_rubles > 0:
                        prices[nmid] = price_rubles
                        read_count += 1
                except (ValueError, TypeError):
                    continue
            except Exception:
                continue
        
        print(f"[OK] Прочитано цен: {read_count}")
        return prices
        
    except Exception as e:
        print(f"[ERROR] Ошибка при чтении Excel файла: {e}")
        import traceback
        traceback.print_exc()
        return prices


def update_prices_via_api(prices_data: List[Dict[str, Any]]) -> bool:
    """
    Обновить цены товаров через API WB
    
    Args:
        prices_data: Список данных о ценах [{"nmID": int, "price": int, "discount": int}]
                    где price в копейках
        
    Returns:
        bool: True если успешно
    """
    url = f"{Config.PRICES_API_URL}/upload/task"
    headers = get_headers()
    
    # Удаляем дубликаты nmID - оставляем последнее значение для каждого nmID
    seen_nmids = {}
    for item in prices_data:
        nmid = item.get("nmID") or item.get("nmId")
        if nmid:
            seen_nmids[int(nmid)] = {
                "nmID": int(nmid),
                "price": int(item["price"]),
                "discount": int(item.get("discount", 0))
            }
    
    data_items = list(seen_nmids.values())
    
    # Если были дубликаты, выводим информацию
    if len(data_items) < len(prices_data):
        print(f"[INFO] Удалено {len(prices_data) - len(data_items)} дубликатов nmID")
    
    if not data_items:
        print("[WARN] Нет данных для обновления (все дубликаты или пустые nmID)")
        return True
    
    payload = {"data": data_items}
    
    try:
        # API требует POST
        response = requests.post(url, headers=headers, json=payload, timeout=120)
        
        # Обрабатываем 429 ошибку (Too Many Requests)
        if response.status_code == 429:
            print("[WARN] Превышен лимит запросов (429), ожидание 5 секунд...")
            time.sleep(5)
            response = requests.post(url, headers=headers, json=payload, timeout=120)
        
        # Обрабатываем 400 ошибки - некоторые не критичны
        if response.status_code == 400:
            try:
                error_data = response.json()
                error_text = error_data.get('errorText', '')
                error_lower = error_text.lower()
                
                if 'already set' in error_lower or 'уже установлены' in error_lower:
                    print("[INFO] Цены уже установлены (не требуют обновления)")
                    return True
                elif 'duplicate' in error_lower:
                    print("[INFO] Обнаружены дубликаты (уже обработано)")
                    return True
            except (ValueError, KeyError):
                pass
        
        response.raise_for_status()
        
        print(f"[OK] Цены успешно обновлены через API ({len(data_items)} товаров)")
        return True
        
    except requests.exceptions.RequestException as e:
        # Проверяем, не является ли это некритичной ошибкой
        if hasattr(e, 'response') and e.response is not None:
            error_text = e.response.text
            error_lower = error_text.lower()
            
            if 'already set' in error_lower or 'уже установлены' in error_lower:
                print("[INFO] Цены уже установлены (не требуют обновления)")
                return True
            elif 'duplicate' in error_lower:
                print("[INFO] Обнаружены дубликаты (уже обработано)")
                return True
        
        print(f"[ERROR] Ошибка при обновлении цен: {e}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"Ответ сервера: {e.response.text}")
        return False


def update_prices_in_batches(prices_dict: Dict[int, int], batch_size: int = 100) -> bool:
    """
    Обновляет цены на WB через API, разбивая на батчи
    
    Args:
        prices_dict: Словарь {nmID: price_in_rubles}
        batch_size: Размер батча (по умолчанию 100)
        
    Returns:
        bool: True если все батчи обработаны успешно
    """
    if not prices_dict:
        print("[WARN] Нет цен для обновления")
        return True
    
    # Преобразуем словарь в список для отправки в API
    prices_data = []
    for nmid, price_rubles in prices_dict.items():
        prices_data.append({
            "nmID": int(nmid),
            "price": int(price_rubles),
            "discount": 0
        })
    
    total_items = len(prices_data)
    total_batches = (total_items + batch_size - 1) // batch_size
    
    print(f"[INFO] Обновление цен через API: {total_items} товаров, {total_batches} батчей")
    
    all_success = True
    
    for i in range(0, total_items, batch_size):
        batch = prices_data[i:i + batch_size]
        batch_num = i // batch_size + 1
        
        print(f"[INFO] Обработка батча {batch_num}/{total_batches} ({len(batch)} товаров)...")
        
        success = update_prices_via_api(batch)
        
        if not success:
            all_success = False
            print(f"[WARN] Батч {batch_num} завершился с ошибкой")
        
        # Задержка между батчами для избежания 429 ошибок
        if i + batch_size < total_items:
            time.sleep(0.5)
    
    return all_success


def main():
    """Основная функция - выполняет все три шага"""
    print("=" * 70)
    print("Обновление цен на Wildberries через Excel шаблон")
    print("=" * 70)
    print()
    
    try:
        Config.validate()
    except ValueError as e:
        print(f"[ERROR] Ошибка конфигурации: {e}")
        return
    
    # Шаг 1: Скачиваем Excel шаблон
    print("[ШАГ 1] Скачивание актуального Excel шаблона с рекомендуемыми ценами...")
    print("-" * 70)
    
    try:
        template_file = download_excel_only()
    except Exception as e:
        print(f"[ERROR] Ошибка при скачивании шаблона: {e}")
        import traceback
        traceback.print_exc()
        return
    
    if not template_file:
        print("[ERROR] Функция download_excel_only() вернула None")
        print("[INFO] Проверьте настройки браузера и cookies")
        # Попробуем найти последний скачанный файл
        found_files = find_wb_template_files(str(Config.TARGET_DIR))
        if found_files:
            template_file = found_files[0]
            print(f"[INFO] Использую найденный файл: {os.path.basename(template_file)}")
        else:
            return
    
    if not isinstance(template_file, str):
        template_file = str(template_file)
    
    if not os.path.exists(template_file):
        print(f"[ERROR] Скачанный файл не найден: {template_file}")
        # Попробуем найти последний скачанный файл
        found_files = find_wb_template_files(str(Config.TARGET_DIR))
        if found_files:
            template_file = found_files[0]
            print(f"[INFO] Использую найденный файл: {os.path.basename(template_file)}")
        else:
            return
    
    print(f"[OK] Шаблон скачан: {os.path.basename(template_file)}")
    print()
    
    # Шаг 2: Корректируем цены (J = N - 1)
    print("[ШАГ 2] Корректировка цен в шаблоне (колонка J = N - 1)...")
    print("-" * 70)
    
    try:
        changes_count = adjust_prices(template_file, verbose=False)
        print(f"[OK] Скорректировано цен: {changes_count}")
    except Exception as e:
        print(f"[ERROR] Ошибка при корректировке цен: {e}")
        import traceback
        traceback.print_exc()
        return
    
    print()
    
    # Шаг 3: Читаем цены из колонки J и загружаем на WB через API
    print("[ШАГ 3] Загрузка скорректированных цен на WB через API...")
    print("-" * 70)
    
    # Читаем цены из колонки J
    prices_dict = read_prices_from_excel_template(template_file)
    
    if not prices_dict:
        print("[ERROR] Не удалось прочитать цены из шаблона")
        return
    
    print(f"[INFO] Прочитано цен для обновления: {len(prices_dict)}")
    print()
    
    # Обновляем цены через API батчами
    success = update_prices_in_batches(prices_dict, batch_size=100)
    
    if success:
        print()
        print("=" * 70)
        print("[SUCCESS] Все шаги выполнены успешно!")
        print(f"  - Шаблон скачан: {os.path.basename(template_file)}")
        print(f"  - Скорректировано цен: {changes_count}")
        print(f"  - Обновлено цен на WB: {len(prices_dict)}")
        print("=" * 70)
    else:
        print()
        print("=" * 70)
        print("[WARN] Обновление завершено с ошибками")
        print("  Проверьте логи выше для деталей")
        print("=" * 70)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[ERROR] Прервано пользователем")
        sys.exit(1)
    except Exception as e:
        print(f"\n[ERROR] Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

