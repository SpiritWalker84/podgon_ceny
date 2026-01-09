#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для обновления остатков и цен товаров на Wildberries по брендам

Автоматическая корректировка цен:
- Получает рекомендуемые цены через API WB
- Автоматически устанавливает цену на 1 рубль меньше рекомендуемой
- Это помогает избежать "непривлекательных цен" на WB
- Если рекомендуемая цена не найдена, используется базовая цена из прайса
"""

import os
import requests
import pandas as pd
from dotenv import load_dotenv
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import csv
import re
import time
from datetime import datetime
import glob
import shutil

# Импортируем функцию корректировки цен из update_prices.py
try:
    from update_prices import adjust_prices, find_wb_template_files
except ImportError:
    # Если модуль не найден, определяем функцию здесь
    import openpyxl
    from openpyxl import load_workbook
    
    def adjust_prices(file_path, column_n=14, column_j=10, verbose=False):
        """Корректирует цены в Excel файле: устанавливает колонку J = N - 1"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл не найден: {file_path}")
        
        wb = load_workbook(file_path)
        ws = wb.active
        max_row = ws.max_row
        changes_count = 0
        
        for row in range(1, max_row + 1):
            cell_n = ws.cell(row=row, column=column_n)
            cell_j = ws.cell(row=row, column=column_j)
            
            if cell_n.value is not None:
                try:
                    value_n = float(cell_n.value)
                    new_value_j = value_n - 1
                    cell_j.value = new_value_j
                    changes_count += 1
                    if verbose:
                        print(f"Строка {row}: N={value_n} -> J={new_value_j}")
                except (ValueError, TypeError):
                    if verbose:
                        print(f"Строка {row}: значение в N не является числом, пропускаем")
        
        wb.save(file_path)
        return changes_count
    
    def find_wb_template_files(directory="."):
        """Находит файлы шаблонов WB по паттерну имени"""
        patterns = [
            "Шаблон обновления цен и скидок*.xlsx",
            "*обновления цен*.xlsx",
            "*WB*.xlsx",
        ]
        found_files = []
        for pattern in patterns:
            found_files.extend(glob.glob(os.path.join(directory, pattern)))
        found_files = list(set(found_files))
        found_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        return found_files

# Загружаем переменные окружения
# Пробуем загрузить из текущей директории и из родительской
load_dotenv()
load_dotenv('.env')  # Явно указываем файл


class Config:
    """Класс для хранения конфигурации из переменных окружения"""
    
    # API настройки
    WB_API_TOKEN: str = os.getenv('WB_API_TOKEN', '')
    STOCKS_API_URL: str = "https://marketplace-api.wildberries.ru/api/v3"
    PRICES_API_URL: str = "https://discounts-prices-api.wildberries.ru/api/v2"
    
    # Пути
    # Используем текущую директорию, если TARGET_DIR не задан в .env
    default_target = str(Path.cwd())  # Текущая директория
    TARGET_DIR: Path = Path(os.getenv('TARGET_DIR', default_target))
    # BASE_DIR - директория с CSV файлами (на сервере обычно ~/wildberries/price)
    # Если не указан, используем TARGET_DIR
    base_dir_env = os.getenv('BASE_DIR', None)
    BASE_DIR: Optional[Path] = Path(base_dir_env) if base_dir_env else None
    
    # Бренды для обработки
    BRANDS: List[str] = ['BOSCH', 'TRIALLI', 'MANN']
    
    # Коэффициент повышения цены
    PRICE_MULTIPLIER: float = 1.5
    
    # Автоматическая корректировка цен (J = N - 1) для избежания непривлекательных цен
    AUTO_ADJUST_PRICES: bool = os.getenv('AUTO_ADJUST_PRICES', 'true').lower() == 'true'
    
    # Автоматическая загрузка Excel шаблона через браузер (требует Selenium)
    AUTO_DOWNLOAD_EXCEL: bool = os.getenv('AUTO_DOWNLOAD_EXCEL', 'true').lower() == 'true'  # По умолчанию включено для теста
    
    # Запускать браузер в видимом режиме (headless=False) - полезно для первой авторизации
    # На Linux сервере обычно нужно true (headless)
    HEADLESS_BROWSER: bool = os.getenv('HEADLESS_BROWSER', 'true').lower() == 'true'
    
    # Путь к Chrome/Edge для Selenium (если не указан, будет использован системный)
    BROWSER_PATH: Optional[str] = os.getenv('BROWSER_PATH', None)
    
    # URL для скачивания шаблона цен WB
    WB_BASE_URL: str = "https://seller.wildberries.ru"
    WB_PRICES_URL: str = "https://seller.wildberries.ru/discount-and-prices"
    
    # Данные для авторизации (если нужно)
    WB_LOGIN: Optional[str] = os.getenv('WB_LOGIN', None)
    WB_PASSWORD: Optional[str] = os.getenv('WB_PASSWORD', None)
    
    # Путь для сохранения cookies (для автоматической авторизации)
    COOKIES_FILE: Path = Path.cwd() / "wb_cookies.pkl"
    
    @classmethod
    def validate(cls) -> None:
        """Проверяет, что все необходимые переменные окружения установлены"""
        if not cls.WB_API_TOKEN:
            raise ValueError("WB_API_TOKEN не установлен в .env файле")


def get_api_token() -> str:
    """
    Получить API токен из .env файла
    
    Returns:
        str: API токен Wildberries
        
    Raises:
        ValueError: Если токен не найден в .env файле
    """
    token = os.getenv('WB_API_TOKEN')
    if not token:
        token = os.getenv('WB_KEY')  # Для обратной совместимости
    
    if not token:
        raise ValueError(
            "API токен не найден в файле .env!\n"
            "Добавьте в .env файл строку: WB_API_TOKEN=ваш_токен"
        )
    
    return token


def get_headers() -> Dict[str, str]:
    """Получить заголовки для API запросов"""
    token = get_api_token()
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }


def get_warehouses() -> List[Dict[str, Any]]:
    """Получить список складов продавца"""
    url = f"{Config.STOCKS_API_URL}/warehouses"
    headers = get_headers()
    
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    
    warehouses = response.json()
    print(f"Найдено складов: {len(warehouses)}")
    for warehouse in warehouses:
        print(f"  - {warehouse.get('name')} (ID: {warehouse.get('id')})")
    
    return warehouses


def read_mapping_files() -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str], Dict[str, str], Dict[str, str]]:
    """
    Читает файл соответствия "Баркоды.xlsx"
    
    Структура файла (данные с 5-й строки):
    - Колонка B (индекс 1) - артикул производителя
    - Колонка C (индекс 2) - nmID (артикул WB)
    - Колонка G (индекс 6) - баркод
    
    Returns:
        Tuple[Dict[str, str], Dict[str, str], Dict[str, str], Dict[str, str], Dict[str, str]]: 
            - Словарь {артикул_производителя: nmID}
            - Словарь {баркод: nmID}
            - Словарь {артикул_производителя: nmID} (дубликат для совместимости)
            - Словарь {артикул_производителя: баркод}
            - Словарь {баркод: chrtId} (пустой, заполняется позже)
    """
    art_to_nmid: Dict[str, str] = {}  # Артикул производителя -> nmID
    barcode_to_nmid: Dict[str, str] = {}  # Баркод -> nmID
    manufacturer_art_to_nmid: Dict[str, str] = {}  # Артикул производителя -> nmID (дубликат)
    manufacturer_art_to_barcode: Dict[str, str] = {}  # Артикул производителя -> баркод
    barcode_to_chrtid: Dict[str, str] = {}
    
    # Ищем файл с баркодами
    barcode_file = None
    for file in os.listdir('.'):
        if 'Баркоды' in file and file.endswith('.xlsx'):
            barcode_file = file
            break
    
    if barcode_file:
        try:
            # Читаем файл, пропуская первые 4 строки (данные начинаются с 5-й строки)
            df_barcode = pd.read_excel(barcode_file, header=0, skiprows=4)
            
            # Структура файла (данные с 5-й строки):
            # Колонка B (индекс 1) - артикул производителя
            # Колонка C (индекс 2) - nmID (артикул WB)
            # Колонка G (индекс 6) - баркод
            if len(df_barcode.columns) >= 7:
                manufacturer_art_col = df_barcode.columns[1]  # Колонка B - артикул производителя
                nmid_col = df_barcode.columns[2]  # Колонка C - nmID
                barcode_col = df_barcode.columns[6]  # Колонка G - баркод
                
                for idx, row in df_barcode.iterrows():
                    try:
                        manufacturer_art = str(row[manufacturer_art_col]).strip() if len(row) > 1 else None
                        nmid_val = row[nmid_col]
                        barcode = str(row[barcode_col]).strip()
                        
                        # Пропускаем заголовки и пустые значения
                        if manufacturer_art.lower() in ['артикул', 'артикул производителя', 'nan', ''] or not manufacturer_art:
                            continue
                        
                        if pd.isna(nmid_val) or not barcode or barcode.lower() in ['баркод', 'barcode', 'баркод в системе', 'nan', ''] or len(barcode) <= 5:
                            continue
                        
                        # Получаем nmID из колонки C
                        try:
                            nmid = str(int(float(nmid_val))).strip()
                        except (ValueError, TypeError):
                            continue
                        
                        if nmid and barcode:
                            # Создаем соответствие артикул производителя -> nmID
                            # Сохраняем все варианты: оригинальный, без пробелов, нормализованный
                            manufacturer_art_clean = manufacturer_art.replace(' ', '').upper()
                            manufacturer_art_normalized = manufacturer_art_clean.replace('-', '').replace('/', '').replace('_', '')
                            
                            art_to_nmid[manufacturer_art] = nmid  # Оригинальный вариант
                            art_to_nmid[manufacturer_art_clean] = nmid  # Без пробелов
                            art_to_nmid[manufacturer_art_normalized] = nmid  # Нормализованный
                            manufacturer_art_to_nmid[manufacturer_art_clean] = nmid
                            
                            # Создаем соответствие баркод -> nmID
                            barcode_to_nmid[barcode] = nmid
                            
                            # Создаем соответствие артикул производителя -> баркод
                            manufacturer_art_to_barcode[manufacturer_art] = barcode  # Оригинальный вариант
                            manufacturer_art_to_barcode[manufacturer_art_clean] = barcode  # Без пробелов
                            manufacturer_art_to_barcode[manufacturer_art_normalized] = barcode  # Нормализованный
                    except (ValueError, TypeError, KeyError, IndexError):
                        continue
        except Exception as e:
            print(f"Ошибка при чтении файла баркодов: {e}")
            import traceback
            traceback.print_exc()
    else:
        print("[WARN] Файл 'Баркоды.xlsx' не найден!")
    
    return art_to_nmid, barcode_to_nmid, manufacturer_art_to_nmid, manufacturer_art_to_barcode, barcode_to_chrtid


def get_chrt_id_by_barcode(barcode: str, warehouse_id: int, stocks_cache: Optional[Dict[str, int]] = None) -> Optional[int]:
    """
    Получить chrtId по баркоду через API или из кэша
    
    Примечание: API может не поддерживать GET с параметром sku, поэтому возвращаем None.
    API сам найдет chrtId по sku при обновлении остатков через PUT запрос.
    
    Args:
        barcode: Баркод товара
        warehouse_id: ID склада
        stocks_cache: Кэш остатков {barcode: chrtId}
        
    Returns:
        Optional[int]: chrtId или None (API найдет автоматически)
    """
    # API не поддерживает получение chrtId через GET запрос с параметром sku
    # Возвращаем None - API сам найдет chrtId по sku при обновлении остатков
    return None


def get_all_stocks(warehouse_id: int) -> Dict[str, int]:
    """
    Получить все остатки со склада и создать кэш {barcode: chrtId}
    
    Примечание: API может не поддерживать получение всех остатков сразу,
    поэтому возвращаем пустой кэш и будем получать chrtId по требованию.
    
    Args:
        warehouse_id: ID склада
        
    Returns:
        Dict[str, int]: Словарь {barcode: chrtId} (обычно пустой)
    """
    # API не поддерживает GET /stocks/{warehouse_id} без параметров
    # Будем получать chrtId по требованию через параметр sku
    return {}


def read_brand_file(brand: str) -> List[Dict[str, Any]]:
    """
    Читает файл бренда и извлекает данные
    
    Args:
        brand: Название бренда
        
    Returns:
        List[Dict[str, Any]]: Список товаров с данными
    """
    # Сначала ищем в BASE_DIR (для сервера: ~/wildberries/price)
    # Если не найдено, ищем в TARGET_DIR
    if Config.BASE_DIR and Path(Config.BASE_DIR).exists():
        brand_file = Path(Config.BASE_DIR) / f"brand_{brand}.csv"
        if not brand_file.exists():
            # Пробуем в TARGET_DIR
            brand_file = Config.TARGET_DIR / f"brand_{brand}.csv"
    else:
        brand_file = Config.TARGET_DIR / f"brand_{brand}.csv"
    
    if not brand_file.exists():
        print(f"  [WARN] Файл не найден: {brand_file}")
        return []
    
    print(f"  [INFO] Читаю файл: {brand_file}")
    
    products = []
    
    # Определяем кодировку и разделитель
    encoding = 'utf-8'
    try:
        with open(brand_file, 'r', encoding='utf-8') as f:
            sample = f.read(1000)
    except UnicodeDecodeError:
        encoding = 'cp1251'
        with open(brand_file, 'r', encoding='cp1251') as f:
            sample = f.read(1000)
    
    # Определяем разделитель
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(sample, delimiters=',;\t')
    except:
        dialect = csv.excel
    
    # Читаем файл
    with open(brand_file, 'r', encoding=encoding) as f:
        reader = csv.reader(f, dialect=dialect)
        
        header = None
        for row_num, row in enumerate(reader):
            if row_num == 0:
                header = row
                continue
            
            if len(row) < 5:
                continue
            
            # Структура файла бренда:
            # Колонка A (0) - бренд
            # Колонка B (1) - возможно артикул продавца или название
            # Колонка C (2) - возможно баркод или другой идентификатор
            # Колонка D (3) - цена
            # Колонка E (4) - количество
            
            try:
                # Извлекаем цену из колонки D (индекс 3)
                price_str = str(row[3]).strip().replace(',', '.').replace(' ', '').replace('"', '')
                # Извлекаем количество из колонки E (индекс 4)
                amount_str = str(row[4]).strip().replace(',', '.').replace(' ', '').replace('"', '')
                
                price = None
                amount = None
                
                if price_str and price_str.lower() not in ['nan', '', 'цена', 'price']:
                    try:
                        price = float(price_str)
                    except ValueError:
                        pass
                
                if amount_str and amount_str.lower() not in ['nan', '', 'количество', 'amount', 'остаток']:
                    try:
                        amount = int(float(amount_str))
                    except ValueError:
                        pass
                
                if price is None or amount is None:
                    continue
                
                # Ищем артикул производителя и баркод
                # В CSV файлах колонка B (индекс 1) содержит артикул производителя (F00BH40270, AG 01007, CUK18000-2)
                # Колонка C (индекс 2) содержит описание товара
                manufacturer_art = None  # Артикул производителя из CSV
                seller_art = None  # Артикул продавца (будет найден через соответствие)
                barcode = None
                
                # Колонка B (индекс 1) - это артикул производителя
                if len(row) > 1 and row[1]:
                    potential_manufacturer_art = str(row[1]).strip().replace('"', '').replace("'", '')
                    # Пропускаем заголовки
                    if (potential_manufacturer_art.lower() not in ['бренд', 'brand', 'артикул', 'артикул продавца', 'название', 'name', 'nan', '', 'none'] and
                        len(potential_manufacturer_art) >= 2 and len(potential_manufacturer_art) <= 20):
                        # Убираем пробелы для сопоставления (AG 01007 -> AG01007)
                        manufacturer_art = potential_manufacturer_art.replace(' ', '')
                
                # Проверяем колонку C (индекс 2) на наличие баркода (маловероятно, но проверим)
                if len(row) > 2 and row[2]:
                    potential_barcode = str(row[2]).strip().replace('"', '').replace("'", '').replace(' ', '').replace('-', '')
                    # Если это длинный баркод (13+ цифр) - EAN-13
                    if len(potential_barcode) >= 13 and potential_barcode.isdigit():
                        barcode = potential_barcode
                
                products.append({
                    'manufacturer_art': manufacturer_art,  # Артикул производителя из CSV
                    'seller_art': seller_art,  # Артикул продавца (будет найден через соответствие)
                    'barcode': barcode,
                    'price': price,
                    'amount': amount,
                    'row': row
                })
            except (ValueError, IndexError, TypeError) as e:
                continue
    
    return products


def update_stocks(warehouse_id: int, stocks_data: List[Dict[str, Any]]) -> bool:
    """
    Обновить остатки на складе
    
    Args:
        warehouse_id: ID склада
        stocks_data: Список данных об остатках [{"chrtId": int, "sku": str, "amount": int}]
        
    Returns:
        bool: True если успешно
    """
    url = f"{Config.STOCKS_API_URL}/stocks/{warehouse_id}"
    headers = get_headers()
    
    payload = {"stocks": stocks_data}
    
    try:
        response = requests.put(url, headers=headers, json=payload, timeout=60)
        
        # Обрабатываем 429 ошибку (Too Many Requests)
        if response.status_code == 429:
            print(f"    [WARN] Превышен лимит запросов (429), ожидание 5 секунд...")
            time.sleep(5)
            # Повторяем запрос после задержки
            response = requests.put(url, headers=headers, json=payload, timeout=60)
        
        response.raise_for_status()
        return True
    except requests.exceptions.RequestException as e:
        print(f"    [ERROR] Ошибка при обновлении остатков: {e}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"    Ответ сервера: {e.response.text}")
        return False


def get_recommended_prices(nmids: List[int]) -> Dict[int, Optional[int]]:
    """
    Получить рекомендуемые цены для товаров через API WB.
    
    Args:
        nmids: Список nmID товаров
    
    Returns:
        Dict[int, Optional[int]]: Словарь {nmID: recommended_price} или None если цена не найдена
    """
    recommended_prices = {}
    
    if not nmids:
        return recommended_prices
    
    print(f"  [INFO] Получение рекомендуемых цен для {len(nmids)} товаров...")
    headers = get_headers()
    
    # Разбиваем на батчи по 100 nmID (лимит API)
    batch_size = 100
    total_batches = (len(nmids) + batch_size - 1) // batch_size
    
    for i in range(0, len(nmids), batch_size):
        batch_nmids = nmids[i:i + batch_size]
        batch_num = i // batch_size + 1
        print(f"  [INFO] Батч {batch_num}/{total_batches} ({len(batch_nmids)} товаров)...")
        
        # Пробуем несколько возможных API endpoints для получения рекомендуемых цен
        # Endpoint 1: GET запрос /info (может содержать рекомендуемые цены)
        try:
            url = f"{Config.PRICES_API_URL}/info"
            params = {"nmIDs": ",".join(map(str, batch_nmids))}
            
            response = requests.get(url, headers=headers, params=params, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                # Обрабатываем ответ в зависимости от формата
                items_to_process = []
                if isinstance(data, list):
                    items_to_process = data
                elif isinstance(data, dict):
                    items_to_process = data.get("data", [])
                
                for item in items_to_process:
                    nmid = item.get("nmID") or item.get("nmId")
                    # Пробуем разные варианты названий полей для рекомендуемой цены
                    recommended_price = (
                        item.get("recommendedPrice") or 
                        item.get("recommended_price") or 
                        item.get("recommendedPriceWithDiscount") or
                        item.get("price") or
                        item.get("minPrice")  # Минимальная рекомендованная цена
                    )
                    if nmid and recommended_price:
                        try:
                            recommended_prices[int(nmid)] = int(recommended_price)
                        except (ValueError, TypeError):
                            pass
        except (requests.exceptions.RequestException, KeyError, ValueError):
            pass
        
        # Endpoint 2: POST запрос /list/goods/filter
        if len(recommended_prices) < len(batch_nmids):
            try:
                url = f"{Config.PRICES_API_URL}/list/goods/filter"
                payload = {"nmIDs": batch_nmids}
                
                response = requests.post(url, headers=headers, json=payload, timeout=10)
                
                if response.status_code == 200:
                    data = response.json()
                    items_to_process = []
                    if isinstance(data, list):
                        items_to_process = data
                    elif isinstance(data, dict):
                        items_to_process = data.get("data", []) or data.get("goods", [])
                    
                    for item in items_to_process:
                        nmid = item.get("nmID") or item.get("nmId")
                        if nmid and nmid not in recommended_prices:
                            recommended_price = (
                                item.get("recommendedPrice") or 
                                item.get("recommended_price") or 
                                item.get("recommendedPriceWithDiscount") or
                                item.get("price") or
                                item.get("minPrice")
                            )
                            if recommended_price:
                                try:
                                    recommended_prices[int(nmid)] = int(recommended_price)
                                except (ValueError, TypeError):
                                    pass
            except (requests.exceptions.RequestException, KeyError, ValueError):
                pass
        
        # Небольшая задержка между батчами
        if i + batch_size < len(nmids):
            time.sleep(0.5)
    
    return recommended_prices


def adjust_price_by_recommended(nmid: int, current_price: int, recommended_prices: Dict[int, Optional[int]]) -> int:
    """
    Корректирует цену: если есть рекомендуемая цена, устанавливает на 1 меньше.
    Иначе возвращает текущую цену.
    
    Args:
        nmid: nmID товара
        current_price: Текущая цена из прайса
        recommended_prices: Словарь рекомендуемых цен {nmID: price}
    
    Returns:
        int: Скорректированная цена
    """
    if nmid in recommended_prices and recommended_prices[nmid] is not None:
        recommended_price = recommended_prices[nmid]
        # Устанавливаем цену на 1 меньше рекомендуемой
        adjusted_price = recommended_price - 1
        # Но не меньше текущей цены * 0.9 (защита от слишком низких цен)
        min_price = int(current_price * 0.9)
        return max(adjusted_price, min_price)
    
    return current_price


def update_prices(prices_data: List[Dict[str, Any]]) -> bool:
    """
    Обновить цены товаров
    
    Args:
        prices_data: Список данных о ценах [{"nmID": int, "price": int, "discount": int}]
        
    Returns:
        bool: True если успешно
    """
    url = f"{Config.PRICES_API_URL}/upload/task"
    headers = get_headers()
    
    # Формируем данные в правильном формате (как в update_prices_stocks_wb.py)
    # Удаляем дубликаты nmID - оставляем последнее значение для каждого nmID
    # Убираем внутренние поля (base_price) перед отправкой в API
    seen_nmids = {}
    for item in prices_data:
        nmid = item.get("nmID") or item.get("nmId")
        if nmid:
            seen_nmids[int(nmid)] = {
                "nmID": int(nmid),
                "price": int(item["price"]),
                "discount": int(item.get("discount", 0))
            }
    
    data_items = list(seen_nmids.values())
    
    # Если были дубликаты, выводим информацию
    if len(data_items) < len(prices_data):
        print(f"    [INFO] Удалено {len(prices_data) - len(data_items)} дубликатов nmID")
    
    if not data_items:
        print(f"    [WARN] Нет данных для обновления (все дубликаты или пустые nmID)")
        return True
    
    payload = {"data": data_items}
    
    try:
        # API требует POST, а не PUT
        response = requests.post(url, headers=headers, json=payload, timeout=120)
        
        # Обрабатываем 429 ошибку (Too Many Requests)
        if response.status_code == 429:
            print(f"    [WARN] Превышен лимит запросов (429), ожидание 5 секунд...")
            time.sleep(5)
            # Повторяем запрос после задержки
            response = requests.post(url, headers=headers, json=payload, timeout=120)
        
        # Обрабатываем 400 ошибки - некоторые не критичны
        if response.status_code == 400:
            try:
                error_data = response.json()
                error_text = error_data.get('errorText', '')
                error_lower = error_text.lower()
                
                if 'already set' in error_lower or 'уже установлены' in error_lower():
                    # Цены уже установлены - это нормально, не считаем ошибкой
                    print(f"    [INFO] Цены уже установлены (не требуют обновления)")
                    return True
                elif 'duplicate' in error_lower:
                    # Дубликаты - это не критично, но лучше их удалить (уже удалены выше)
                    print(f"    [INFO] Обнаружены дубликаты (уже обработано)")
                    return True
            except (ValueError, KeyError):
                pass
        
        response.raise_for_status()
        return True
    except requests.exceptions.RequestException as e:
        # Проверяем, не является ли это некритичной ошибкой
        if hasattr(e, 'response') and e.response is not None:
            error_text = e.response.text
            error_lower = error_text.lower()
            
            if 'already set' in error_lower or 'уже установлены' in error_lower:
                # Цены уже установлены - это нормально, не считаем ошибкой
                print(f"    ℹ Цены уже установлены (не требуют обновления)")
                return True
            elif 'duplicate' in error_lower:
                # Дубликаты - это не критично (должны быть удалены выше, но на всякий случай)
                print(f"    ℹ Обнаружены дубликаты (уже обработано)")
                return True
        
        # Это настоящая ошибка
        print(f"    [ERROR] Ошибка при обновлении цен: {e}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"    Ответ сервера: {e.response.text}")
        return False


def read_prices_from_excel_template() -> Dict[int, int]:
    """
    Читает рекомендуемые цены (колонка N) из Excel шаблона WB.
    
    Returns:
        Dict[int, int]: Словарь {nmID: recommended_price} из колонки N
    """
    recommended_prices = {}
    
    # Ищем Excel файл шаблона
    found_files = find_wb_template_files(str(Config.TARGET_DIR))
    
    if not found_files:
        print("  [WARN] Excel шаблон не найден, рекомендуемые цены не будут использованы")
        return recommended_prices
    
    template_file = found_files[0]
    print(f"  [INFO] Читаю рекомендуемые цены из: {os.path.basename(template_file)}")
    
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        wb = load_workbook(template_file, data_only=True)  # data_only=True для получения значений формул
        ws = wb.active
        
        # Ищем колонки:
        # Колонка C (3) - nmID
        # Колонка N (14) - рекомендуемая цена
        # Колонка J (10) - текущая цена (может быть скорректирована)
        
        header_row = 1
        nmid_col = 3  # Колонка C
        recommended_price_col = 14  # Колонка N
        
        # Читаем данные начиная со второй строки
        for row in range(2, ws.max_row + 1):
            try:
                nmid_cell = ws.cell(row=row, column=nmid_col)
                price_cell = ws.cell(row=row, column=recommended_price_col)
                
                if nmid_cell.value and price_cell.value:
                    try:
                        nmid = int(float(str(nmid_cell.value)))
                        recommended_price = int(float(str(price_cell.value)))
                        recommended_prices[nmid] = recommended_price
                    except (ValueError, TypeError):
                        continue
            except Exception:
                continue
        
        print(f"  [OK] Прочитано рекомендуемых цен: {len(recommended_prices)}")
        return recommended_prices
        
    except Exception as e:
        print(f"  [ERROR] Ошибка при чтении Excel шаблона: {e}")
        return recommended_prices


def download_excel_via_cookies() -> Optional[str]:
    """
    Пытается скачать Excel шаблон используя сохраненные cookies через requests.
    Работает на Linux сервере без браузера.
    
    Returns:
        Optional[str]: Путь к скачанному файлу или None если не удалось
    """
    print("  [INFO] Пробую загрузить через сохраненные cookies...")
    
    # Пробуем использовать сохраненные cookies
    cookies_file = Config.COOKIES_FILE
    
    if not cookies_file.exists():
        print("  [INFO] Файл cookies не найден, пропускаю")
        return None
    
    try:
        import pickle
        print("[INFO] Использую сохраненные cookies для загрузки Excel...")
        
        with open(cookies_file, 'rb') as f:
            cookies_list = pickle.load(f)
        
        # Преобразуем список cookies в словарь для requests
        cookies_dict = {}
        for cookie in cookies_list:
            if 'name' in cookie and 'value' in cookie:
                cookies_dict[cookie['name']] = cookie['value']
        
        if not cookies_dict:
            return None
        
        # Используем session для сохранения cookies
        session = requests.Session()
        session.cookies.update(cookies_dict)
        
        # Добавляем заголовки
        headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36',
            'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/vnd.ms-excel, */*',
        }
        
        download_dir = str(Config.TARGET_DIR)
        os.makedirs(download_dir, exist_ok=True)
        
        # Пробуем различные возможные endpoints
        base_url = Config.WB_BASE_URL
        endpoints = [
            f"{base_url}/discount-and-prices/export/template",
            f"{base_url}/api/v2/prices/template/download",
            f"{Config.PRICES_API_URL}/template/download",
        ]
        
        for endpoint in endpoints:
            try:
                print(f"  [INFO] Пробую: {endpoint}")
                response = session.get(endpoint, headers=headers, timeout=5, stream=True, allow_redirects=True)  # Уменьшил таймаут
                
                if response.status_code == 200:
                    content_type = response.headers.get('Content-Type', '')
                    
                    # Проверяем, что это Excel файл
                    if 'excel' in content_type.lower() or 'spreadsheet' in content_type.lower() or 'application/vnd.openxmlformats' in content_type:
                        filename = f"Шаблон обновления цен и скидок {datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        
                        # Пытаемся получить имя файла из заголовка
                        content_disposition = response.headers.get('Content-Disposition', '')
                        if content_disposition:
                            import re
                            match = re.search(r'filename[^;=\n]*=(([\'"]).*?\2|[^;\n]*)', content_disposition)
                            if match:
                                filename = match.group(1).strip('"\'').replace(';', '').strip()
                        
                        file_path = Path(download_dir) / filename
                        
                        # Сохраняем файл
                        with open(file_path, 'wb') as f:
                            for chunk in response.iter_content(chunk_size=8192):
                                f.write(chunk)
                        
                        print(f"  [OK] Файл скачан через cookies: {filename}")
                        return str(file_path)
            except requests.exceptions.RequestException:
                continue
        
        return None
    except Exception as e:
        print(f"  [WARN] Ошибка при использовании cookies: {e}")
        return None


def download_excel_via_api() -> Optional[str]:
    """
    Пытается скачать Excel шаблон напрямую через API/HTTP запрос.
    Использует API токен для авторизации.
    
    Returns:
        Optional[str]: Путь к скачанному файлу или None если не удалось
    """
    print("[INFO] Попытка загрузки Excel шаблона через API...")
    
    headers = get_headers()
    download_dir = str(Config.TARGET_DIR)
    os.makedirs(download_dir, exist_ok=True)
    
    # Пробуем различные возможные endpoints для скачивания шаблона
    api_endpoints = [
        f"{Config.PRICES_API_URL}/template/download",
        f"{Config.PRICES_API_URL}/template",
        f"{Config.PRICES_API_URL}/export/template",
        "https://seller.wildberries.ru/api/v2/prices/template",
    ]
    
    for endpoint in api_endpoints:
        try:
            print(f"  [INFO] Пробую endpoint: {endpoint}")
            response = requests.get(endpoint, headers=headers, timeout=5, stream=True)  # Уменьшил таймаут до 5 сек
            
            if response.status_code == 200:
                # Проверяем, что это Excel файл
                content_type = response.headers.get('Content-Type', '')
                content_disposition = response.headers.get('Content-Disposition', '')
                
                if 'excel' in content_type.lower() or 'spreadsheet' in content_type.lower() or '.xlsx' in content_disposition.lower():
                    # Генерируем имя файла
                    filename = "Шаблон обновления цен и скидок.xlsx"
                    if content_disposition:
                        # Пытаемся извлечь имя файла из заголовка
                        import re
                        match = re.search(r'filename[^;=\n]*=(([\'"]).*?\2|[^;\n]*)', content_disposition)
                        if match:
                            filename = match.group(1).strip('"\'')
                    
                    file_path = Path(download_dir) / filename
                    
                    # Сохраняем файл
                    with open(file_path, 'wb') as f:
                        for chunk in response.iter_content(chunk_size=8192):
                            f.write(chunk)
                    
                    print(f"  [OK] Файл скачан через API: {filename}")
                    return str(file_path)
        except requests.exceptions.RequestException as e:
            continue
    
    print("  [WARN] Не удалось скачать через API, пробую браузерную автоматизацию...")
    return None


def download_excel_template_automated() -> Optional[str]:
    """
    Автоматически скачивает Excel шаблон с рекомендуемыми ценами.
    Сначала пытается через API, если не получается - через браузерную автоматизацию.
    
    Returns:
        Optional[str]: Путь к скачанному файлу или None если не удалось
    """
    if not Config.AUTO_DOWNLOAD_EXCEL:
        print("[INFO] AUTO_DOWNLOAD_EXCEL отключен, пропускаю загрузку")
        return None
    
    print("[INFO] Начинаю попытки скачать Excel шаблон...")
    print(f"[DEBUG] TARGET_DIR: {Config.TARGET_DIR}")
    
    # Приоритет методов скачивания (от простого к сложному):
    # 1. Через сохраненные cookies + requests (работает на Linux без браузера)
    print("\n[INFO] Метод 1: Через сохраненные cookies...")
    try:
        file_path = download_excel_via_cookies()
        if file_path:
            print(f"[SUCCESS] Файл скачан через cookies: {file_path}")
            return file_path
        print("  [INFO] Метод 1 не сработал")
    except Exception as e:
        print(f"  [ERROR] Ошибка в методе 1: {e}")
    
    # 2. Через API напрямую (быстрее и работает на Linux без GUI)
    print("\n[INFO] Метод 2: Через API...")
    try:
        file_path = download_excel_via_api()
        if file_path:
            print(f"[SUCCESS] Файл скачан через API: {file_path}")
            return file_path
        print("  [INFO] Метод 2 не сработал")
    except Exception as e:
        print(f"  [ERROR] Ошибка в методе 2: {e}")
    
    # 3. Через браузерную автоматизацию
    print("\n[INFO] Метод 3: Через браузерную автоматизацию...")
    print("[INFO] Запускаю браузер... (это может занять время)")
    
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.edge.service import Service as EdgeService
        from selenium.webdriver.edge.options import Options as EdgeOptions
    except ImportError:
        print("  [ERROR] Selenium не установлен. Установите: py -m pip install selenium")
        print("  [INFO] Автоматическая загрузка Excel отключена")
        return None
    
    download_dir = str(Config.TARGET_DIR)
    os.makedirs(download_dir, exist_ok=True)
    
    # Настройка браузера
    browser_options = None
    service = None
    driver = None
    
    try:
        # Пробуем использовать Chrome/Chromium (работает на Linux с headless)
        try:
            chrome_options = Options()
            if Config.HEADLESS_BROWSER:
                chrome_options.add_argument("--headless")  # Без интерфейса (обязательно для Linux)
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")  # Важно для Linux серверов
            chrome_options.add_argument("--disable-dev-shm-usage")  # Важно для Linux
            chrome_options.add_argument("--remote-debugging-port=9222")  # Для отладки
            
            # Настройка директории для скачивания
            prefs = {
                "download.default_directory": download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True
            }
            chrome_options.add_experimental_option("prefs", prefs)
            
            if Config.BROWSER_PATH:
                chrome_options.binary_location = Config.BROWSER_PATH
            
            print("  [DEBUG] Создаю Chrome WebDriver...")
            driver = webdriver.Chrome(options=chrome_options)
            print("  [OK] Chrome/Chromium браузер запущен успешно")
            print(f"  [DEBUG] Текущий URL: {driver.current_url}")
        except Exception as chrome_error:
            print(f"  [WARN] Chrome не запустился: {chrome_error}")
            # Пробуем Edge (обычно установлен в Windows)
            print("  [INFO] Пробую запустить Edge...")
            try:
                edge_options = EdgeOptions()
                if Config.HEADLESS_BROWSER:
                    edge_options.add_argument("--headless")  # Без интерфейса
                edge_options.add_argument("--disable-gpu")
                edge_options.add_argument("--no-sandbox")
                
                # Настройка директории для скачивания
                prefs = {
                    "download.default_directory": download_dir,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "safebrowsing.enabled": True
                }
                edge_options.add_experimental_option("prefs", prefs)
                
                if Config.BROWSER_PATH:
                    edge_options.binary_location = Config.BROWSER_PATH
                
                driver = webdriver.Edge(options=edge_options)
                print("  [OK] Edge браузер запущен успешно")
            except Exception as edge_error:
                print(f"  [ERROR] Не удалось запустить браузер: {chrome_error}, {edge_error}")
                print("  [INFO] На Linux сервере установите: sudo apt-get install chromium-browser chromium-chromedriver")
                return None
        
        # Загружаем сохраненные cookies если есть
        driver.get(Config.WB_BASE_URL)
        time.sleep(2)
        
        if Config.COOKIES_FILE.exists():
            try:
                import pickle
                print("  [INFO] Загружаю сохраненные cookies...")
                with open(Config.COOKIES_FILE, 'rb') as f:
                    cookies = pickle.load(f)
                    for cookie in cookies:
                        try:
                            driver.add_cookie(cookie)
                        except:
                            pass
                driver.refresh()
                time.sleep(3)
                print("  [OK] Cookies загружены")
            except Exception as e:
                print(f"  [WARN] Не удалось загрузить cookies: {e}")
        
        # Проверяем, авторизованы ли мы (проверяем наличие элементов личного кабинета)
        is_authorized = False
        try:
            # Проверяем наличие элементов, которые есть только в авторизованном кабинете
            # Ищем элементы продавца (меню, кнопки управления товарами и т.д.)
            WebDriverWait(driver, 5).until(
                lambda d: "seller.wildberries.ru" in d.current_url and 
                         ("login" not in d.current_url.lower() and "auth" not in d.current_url.lower())
            )
            # Дополнительная проверка - ищем типичные элементы кабинета продавца
            seller_elements = driver.find_elements(By.XPATH, "//*[contains(text(), 'Товары') or contains(text(), 'Аналитика') or contains(text(), 'Продажи')]")
            if seller_elements:
                is_authorized = True
                print("  [OK] Авторизация подтверждена (найдены элементы кабинета продавца)")
            else:
                # Проверяем URL - если мы на главной странице seller.wildberries.ru, но не на login/auth
                current_url = driver.current_url.lower()
                if "seller.wildberries.ru" in current_url and "login" not in current_url and "auth" not in current_url:
                    is_authorized = True
                    print("  [OK] Авторизация подтверждена (URL указывает на кабинет продавца)")
        except:
            # Если не удалось подтвердить авторизацию, проверяем наличие полей авторизации
            try:
                login_input = driver.find_element(By.NAME, "phone")
                is_authorized = False
                print("  [WARN] Обнаружены поля авторизации - требуется вход")
            except:
                # Если нет полей авторизации и не на странице login/auth, возможно уже авторизованы
                current_url = driver.current_url.lower()
                if "login" not in current_url and "auth" not in current_url:
                    is_authorized = True
                    print("  [INFO] Авторизация предположительно активна (нет полей входа)")
        
        needs_auth = not is_authorized
        
        if needs_auth:
            if Config.WB_LOGIN and Config.WB_PASSWORD:
                print("  [INFO] Выполняю авторизацию...")
                try:
                    login_input = driver.find_element(By.NAME, "phone")
                    login_input.send_keys(Config.WB_LOGIN)
                    time.sleep(1)
                    
                    # Ищем поле пароля
                    password_input = driver.find_element(By.NAME, "password")
                    password_input.send_keys(Config.WB_PASSWORD)
                    time.sleep(1)
                    
                    # Нажимаем кнопку входа
                    login_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Войти') or contains(text(), 'Вход')]")
                    login_button.click()
                    time.sleep(5)  # Ждем авторизации
                    
                    # Ждем перехода или закрытия формы авторизации
                    try:
                        WebDriverWait(driver, 10).until(
                            lambda d: "seller.wildberries.ru" in d.current_url and "login" not in d.current_url.lower()
                        )
                    except:
                        pass
                    
                    # Сохраняем cookies после успешной авторизации
                    try:
                        import pickle
                        with open(Config.COOKIES_FILE, 'wb') as f:
                            pickle.dump(driver.get_cookies(), f)
                        print("  [OK] Cookies сохранены для следующего запуска")
                    except Exception as e:
                        print(f"  [WARN] Не удалось сохранить cookies: {e}")
                        
                except Exception as e:
                    print(f"  [ERROR] Ошибка авторизации: {e}")
                    print("  [WARN] Продолжаю без авторизации")
            else:
                print("  [WARN] Требуется авторизация, но логин/пароль не указаны в .env")
                print("  [INFO] Добавьте в .env: WB_LOGIN=ваш_телефон и WB_PASSWORD=ваш_пароль")
                print("  [INFO] Или откройте браузер вручную и авторизуйтесь перед запуском скрипта")
        else:
            print("  [INFO] Авторизация не требуется (cookies валидны или уже авторизованы)")
        
        # Переходим на страницу цен и скидок
        print(f"  [INFO] Перехожу на страницу: {Config.WB_PRICES_URL}")
        driver.get(Config.WB_PRICES_URL)
        print(f"  [DEBUG] Текущий URL: {driver.current_url}")
        print(f"  [DEBUG] Заголовок страницы: {driver.title}")
        time.sleep(3)
        
        # Проверяем, не на странице ли авторизации (после перехода на страницу цен)
        current_url_after = driver.current_url.lower()
        if "auth" in current_url_after or "login" in current_url_after:
            print("  [WARN] После перехода обнаружена страница авторизации!")
            
            # Если cookies были загружены, но не помогли - пробуем еще раз
            if Config.COOKIES_FILE.exists():
                print("  [INFO] Cookies были загружены, но не сработали - пробую перезагрузить страницу...")
                driver.refresh()
                time.sleep(3)
                current_url_after = driver.current_url.lower()
            
            # Если все еще на странице авторизации, пробуем автоматическую авторизацию
            if "auth" in current_url_after or "login" in current_url_after:
                # Если есть логин/пароль - используем автоматическую авторизацию
                if Config.WB_LOGIN and Config.WB_PASSWORD:
                    print("  [INFO] Пробую автоматическую авторизацию через логин/пароль...")
                    try:
                        login_input = driver.find_element(By.NAME, "phone")
                        login_input.clear()
                        login_input.send_keys(Config.WB_LOGIN)
                        time.sleep(0.5)
                        
                        password_input = driver.find_element(By.NAME, "password")
                        password_input.clear()
                        password_input.send_keys(Config.WB_PASSWORD)
                        time.sleep(0.5)
                        
                        login_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Войти') or contains(text(), 'Вход')]")
                        login_button.click()
                        time.sleep(5)
                        
                        # Проверяем успешность авторизации
                        try:
                            WebDriverWait(driver, 10).until(
                                lambda d: "seller.wildberries.ru" in d.current_url and "login" not in d.current_url.lower()
                            )
                            print("  [OK] Автоматическая авторизация успешна!")
                            # Сохраняем cookies
                            try:
                                import pickle
                                with open(Config.COOKIES_FILE, 'wb') as f:
                                    pickle.dump(driver.get_cookies(), f)
                                print("  [OK] Cookies сохранены!")
                            except Exception as e:
                                print(f"  [WARN] Не удалось сохранить cookies: {e}")
                            
                            # Переходим на нужную страницу
                            print(f"  [INFO] Перехожу на страницу: {Config.WB_PRICES_URL}")
                            driver.get(Config.WB_PRICES_URL)
                            time.sleep(3)
                        except:
                            print("  [WARN] Автоматическая авторизация не удалась")
                    except Exception as e:
                        print(f"  [WARN] Ошибка автоматической авторизации: {e}")
                
                # Если headless режим выключен и авторизация не прошла - даем возможность вручную
                if ("auth" in driver.current_url.lower() or "login" in driver.current_url.lower()) and not Config.HEADLESS_BROWSER:
                    print("  [INFO] Браузер открыт - вы можете авторизоваться вручную")
                    print("  [INFO] Ожидаю максимум 15 секунд...")
                    
                    max_wait_auth = 15  # Сократил время ожидания
                    waited_auth = 0
                    while waited_auth < max_wait_auth:
                        time.sleep(2)
                        waited_auth += 2
                        current_url = driver.current_url.lower()
                        
                        if "auth" not in current_url and "login" not in current_url:
                            print("  [OK] Авторизация обнаружена!")
                            try:
                                import pickle
                                with open(Config.COOKIES_FILE, 'wb') as f:
                                    pickle.dump(driver.get_cookies(), f)
                                print("  [OK] Cookies сохранены!")
                            except Exception as e:
                                print(f"  [WARN] Не удалось сохранить cookies: {e}")
                            
                            print(f"  [INFO] Перехожу на страницу: {Config.WB_PRICES_URL}")
                            driver.get(Config.WB_PRICES_URL)
                            time.sleep(3)
                            break
                        
                        if waited_auth % 5 == 0:
                            print(f"  [INFO] Ожидание... ({waited_auth}/{max_wait_auth} сек)")
                    
                    # Если все еще на странице авторизации после ожидания - выходим
                    if "auth" in driver.current_url.lower() or "login" in driver.current_url.lower():
                        print("  [ERROR] Авторизация не завершена - невозможно продолжить")
                        if driver:
                            driver.quit()
                        return None
                elif ("auth" in current_url_after or "login" in current_url_after) and Config.HEADLESS_BROWSER:
                    print("  [ERROR] Требуется авторизация, но браузер в headless режиме")
                    print("  [INFO] Установите HEADLESS_BROWSER=false в .env или добавьте WB_LOGIN и WB_PASSWORD")
                    if driver:
                        driver.quit()
                    return None
        
        # Ищем элементы согласно структуре:
        # 1. Вкладка/меню "Товары и цены" или "Цены и скидки"
        # 2. Выпадающее меню "Цены и скидки"
        # 3. Кнопка "Обновить через Excel"
        print("  [INFO] Ищу элементы на странице...")
        time.sleep(2)  # Даем странице загрузиться
        
        excel_button = None
        
        # Шаг 1: Ищем меню/кнопку "Цены и скидки" или "Товары и цены"
        print("  [INFO] Шаг 1: Ищу меню 'Цены и скидки'...")
        menu_selectors = [
            "//*[contains(text(), 'Цены и скидки')]",
            "//*[contains(text(), 'Товары и цены')]",
            "//button[contains(., 'Цены')]",
            "//a[contains(., 'Цены')]",
            "//*[@role='button' or @role='menuitem'][contains(., 'Цены')]",
        ]
        
        menu_element = None
        for selector in menu_selectors:
            try:
                elements = driver.find_elements(By.XPATH, selector)
                for elem in elements:
                    if elem.is_displayed():
                        menu_element = elem
                        print(f"  [OK] Найдено меню: '{elem.text[:50]}'")
                        break
                if menu_element:
                    break
            except:
                continue
        
        # Если нашли меню, пробуем кликнуть (может открыть выпадающее меню)
        if menu_element:
            try:
                driver.execute_script("arguments[0].scrollIntoView(true);", menu_element)
                time.sleep(0.5)
                menu_element.click()
                print("  [OK] Меню открыто")
                time.sleep(2)
            except:
                pass
        
        # Шаг 2: Ищем кнопку "Обновить через Excel" в выпадающем меню или на странице
        print("  [INFO] Шаг 2: Ищу кнопку 'Обновить через Excel'...")
        excel_selectors = [
            "//*[contains(text(), 'Обновить через Excel')]",
            "//*[contains(text(), 'Excel') and contains(text(), 'Обновить')]",
            "//button[contains(., 'Excel')]",
            "//a[contains(., 'Excel')]",
            "//*[@role='button' or @role='menuitem'][contains(., 'Excel')]",
        ]
        
        for selector in excel_selectors:
            try:
                elements = driver.find_elements(By.XPATH, selector)
                for elem in elements:
                    text = elem.text.lower()
                    if "excel" in text and elem.is_displayed() and elem.is_enabled():
                        excel_button = elem
                        print(f"  [OK] Найдена кнопка: '{elem.text[:50]}'")
                        break
                if excel_button:
                    break
            except:
                continue
        
        if excel_button:
            print("  [INFO] Кликаю на кнопку 'Обновить через Excel'...")
            try:
                driver.execute_script("arguments[0].scrollIntoView(true);", excel_button)
                time.sleep(1)
                excel_button.click()
                print("  [OK] Кнопка нажата")
                time.sleep(3)  # Ждем открытия окна
            except Exception as e:
                print(f"  [WARN] Ошибка при клике: {e}, пробую через JavaScript")
                try:
                    driver.execute_script("arguments[0].click();", excel_button)
                    time.sleep(3)
                except:
                    pass
        else:
            print("  [WARN] Кнопка 'Обновить через Excel' не найдена")
            print("  [INFO] Возможно структура страницы отличается")
            if driver:
                driver.quit()
            return None
        
        # Шаг 3: В появившемся модальном окне ищем кнопку "Сформировать шаблон"
        if excel_button:
            print("  [INFO] Шаг 3: Ищу кнопку 'Сформировать шаблон' в модальном окне...")
            time.sleep(2)  # Ждем открытия модального окна
            
            create_template_selectors = [
                "//button[contains(text(), 'Сформировать шаблон')]",
                "//button[contains(text(), 'Сформировать')]",
                "//*[contains(text(), 'Сформировать шаблон')]",
                "//*[@role='button' or @type='button'][contains(., 'Сформировать')]",
            ]
            
            create_button = None
            # Пробуем найти в течение 10 секунд
            for attempt in range(5):
                for selector in create_template_selectors:
                    try:
                        elements = driver.find_elements(By.XPATH, selector)
                        for elem in elements:
                            if elem.is_displayed() and elem.is_enabled():
                                create_button = elem
                                print(f"  [OK] Найдена кнопка: '{elem.text[:50]}'")
                                break
                        if create_button:
                            break
                    except:
                        continue
                if create_button:
                    break
                time.sleep(2)
            
            if create_button:
                print("  [INFO] Нажимаю кнопку 'Сформировать шаблон'...")
                try:
                    driver.execute_script("arguments[0].scrollIntoView(true);", create_button)
                    time.sleep(0.5)
                    create_button.click()
                    print("  [OK] Кнопка нажата, начинаю формирование шаблона...")
                    time.sleep(2)
                except Exception as e:
                    print(f"  [WARN] Ошибка при клике: {e}, пробую через JavaScript")
                    try:
                        driver.execute_script("arguments[0].click();", create_button)
                        time.sleep(2)
                    except:
                        pass
                
                # Шаг 4: Ждем формирования шаблона и ищем кнопку "Скачать шаблон"
                if create_button:
                    print("  [INFO] Шаг 4: Ожидаю формирования шаблона и кнопку 'Скачать шаблон'...")
                    download_button = None
                    max_wait_for_ready = 90  # Максимум 90 секунд на формирование
                    waited = 0
                    
                    while waited < max_wait_for_ready:
                        download_selectors = [
                            "//button[contains(text(), 'Скачать шаблон')]",
                            "//button[contains(text(), 'Скачать')]",
                            "//a[contains(text(), 'Скачать шаблон')]",
                            "//a[contains(text(), 'Скачать')]",
                            "//*[@role='button'][contains(., 'Скачать')]",
                            "//*[contains(@class, 'download')]//button",
                            "//*[contains(@class, 'download')]//a",
                        ]
                        
                        for selector in download_selectors:
                            try:
                                elements = driver.find_elements(By.XPATH, selector)
                                for elem in elements:
                                    text = elem.text.lower()
                                    if "скачать" in text and elem.is_displayed() and elem.is_enabled():
                                        download_button = elem
                                        print(f"  [OK] Кнопка 'Скачать шаблон' найдена: '{elem.text[:50]}'")
                                        break
                                if download_button:
                                    break
                            except:
                                continue
                        
                        if download_button:
                            break
                        
                        time.sleep(2)
                        waited += 2
                        
                        if waited % 10 == 0:
                            print(f"  [INFO] Ожидание формирования шаблона... ({waited}/{max_wait_for_ready} сек)")
                    
                    if download_button:
                        print("  [INFO] Нажимаю кнопку 'Скачать шаблон'...")
                        try:
                            driver.execute_script("arguments[0].scrollIntoView(true);", download_button)
                            time.sleep(0.5)
                            download_button.click()
                            print("  [OK] Кнопка скачивания нажата")
                            time.sleep(3)  # Ждем начала скачивания
                        except Exception as e:
                            print(f"  [WARN] Ошибка при клике: {e}, пробую через JavaScript")
                            try:
                                driver.execute_script("arguments[0].click();", download_button)
                                time.sleep(3)
                            except:
                                pass
                        
                        # Ждем завершения скачивания (проверяем наличие нового файла)
                        max_wait = 30
                        waited = 0
                        while waited < max_wait:
                            # Ищем новые Excel файлы в директории загрузок
                            excel_files = list(Path(download_dir).glob("*.xlsx"))
                            # Берем самый новый файл
                            if excel_files:
                                newest_file = max(excel_files, key=os.path.getmtime)
                                # Проверяем, что файл не старый (создан в последние 60 секунд)
                                if time.time() - os.path.getmtime(newest_file) < 60:
                                    print(f"  [OK] Файл скачан: {newest_file.name}")
                                    return str(newest_file)
                            time.sleep(1)
                            waited += 1
                        
                        print("  [WARN] Файл не найден после скачивания")
                    else:
                        print("  [WARN] Кнопка 'Скачать шаблон' не найдена или еще не активна")
            else:
                print("  [WARN] Кнопка 'Сформировать шаблон' не найдена")
        
    except Exception as e:
        print(f"  [ERROR] Ошибка при скачивании через браузер: {e}")
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass
    
    return None


def auto_adjust_wb_template_prices() -> None:
    """
    Автоматически корректирует цены в Excel файлах шаблонов WB.
    Находит все файлы шаблонов и устанавливает J = N - 1 для избежания
    "непривлекательных цен" на Wildberries.
    """
    print("[INFO] Поиск Excel файлов шаблонов WB для автоматической корректировки цен...")
    
    # Ищем в текущей директории и в BASE_DIR
    search_dirs = [os.getcwd()]
    if hasattr(Config, 'BASE_DIR') and Config.BASE_DIR.exists():
        search_dirs.append(str(Config.BASE_DIR))
    if hasattr(Config, 'TARGET_DIR') and Config.TARGET_DIR.exists():
        search_dirs.append(str(Config.TARGET_DIR))
    
    total_adjusted = 0
    for search_dir in search_dirs:
        found_files = find_wb_template_files(search_dir)
        
        for file_path in found_files:
            try:
                print(f"  [INFO] Обработка: {os.path.basename(file_path)}")
                changes_count = adjust_prices(file_path, verbose=False)
                if changes_count > 0:
                    print(f"    [OK] Скорректировано {changes_count} цен")
                    total_adjusted += changes_count
                else:
                    print(f"    [INFO] Изменений не требуется")
            except Exception as e:
                print(f"    [ERROR] Ошибка при обработке файла: {e}")
    
    if total_adjusted > 0:
        print(f"[OK] Автоматическая корректировка завершена. Всего скорректировано цен: {total_adjusted}\n")
    else:
        print("[INFO] Excel файлы шаблонов не найдены или не требуют корректировки\n")


def main() -> None:
    """Основная функция"""
    try:
        Config.validate()
    except ValueError as e:
        print(f"Ошибка конфигурации: {e}")
        return
    
    # Получаем список складов
    try:
        warehouses = get_warehouses()
    except requests.exceptions.RequestException as e:
        print(f"Ошибка при получении списка складов: {e}")
        return
    
    if not warehouses:
        print("Ошибка: не найдено складов")
        return
    
    # Читаем файлы соответствия
    print(f"\n[INFO] Ищу файлы в директории: {Config.TARGET_DIR}")
    art_to_nmid, barcode_to_nmid, manufacturer_art_to_nmid, manufacturer_art_to_barcode, barcode_to_chrtid = read_mapping_files()
    
    if not art_to_nmid and not barcode_to_nmid:
        print("[WARN] Предупреждение: не найдено файлов соответствия")
    else:
        print(f"[INFO] Загружено соответствий: артикулов={len(art_to_nmid)}, баркодов={len(barcode_to_nmid)}")
    
    # Обрабатываем каждый бренд
    print(f"\n[INFO] Обработка брендов: {Config.BRANDS}")
    
    all_stocks_data: Dict[int, List[Dict[str, Any]]] = {}  # {warehouse_id: [stocks]}
    all_prices_data: List[Dict[str, Any]] = []
    
    for brand in Config.BRANDS:
        print(f"\n[INFO] Обработка бренда: {brand}")
        products = read_brand_file(brand)
        print(f"  [INFO] Загружено товаров из файла: {len(products)}")
        
        if not products:
            continue
        
        matched_count = 0
        processed_count = 0
        
        for product in products:
            processed_count += 1
            if processed_count % 100 == 0:
                print(f"  [INFO] Обработано товаров: {processed_count}/{len(products)}")
            nmid = None
            
            # Проверяем только артикулы, которые есть в файле "Баркоды.xlsx"
            # Если артикула нет в файле соответствия, пропускаем товар
            if not product.get('manufacturer_art'):
                continue
            
            manufacturer_art = str(product['manufacturer_art']).strip()
            manufacturer_art_clean = manufacturer_art.replace(' ', '').upper()
            manufacturer_art_normalized = manufacturer_art_clean.replace('-', '').replace('/', '').replace('_', '')
            
            # Проверяем, есть ли артикул в файле соответствия
            art_found = False
            if manufacturer_art in art_to_nmid:
                nmid = art_to_nmid[manufacturer_art]
                art_found = True
            elif manufacturer_art_clean in art_to_nmid:
                nmid = art_to_nmid[manufacturer_art_clean]
                art_found = True
            else:
                # Пробуем найти с учетом нормализации (убираем дефисы, слэши и т.д.)
                for art_key, art_nmid in art_to_nmid.items():
                    art_key_normalized = str(art_key).strip().replace(' ', '').upper().replace('-', '').replace('/', '').replace('_', '')
                    if art_key_normalized == manufacturer_art_normalized:
                        nmid = art_nmid
                        art_found = True
                        break
            
            # Если артикул не найден в файле соответствия, пропускаем товар
            # (это означает, что карточка еще не создана на WB)
            if not art_found:
                continue
            
            matched_count += 1
            
            # Подготавливаем данные для обновления цен
            # Сохраняем исходную цену для дальнейшей корректировки
            base_price = int(product['price'] * Config.PRICE_MULTIPLIER)
            all_prices_data.append({
                "nmID": int(nmid),
                "price": base_price,  # Временно сохраняем базовую цену, скорректируем позже
                "base_price": base_price,  # Сохраняем для корректировки
                "discount": 0
            })
            
            # Подготавливаем данные для обновления остатков
            # Обновляем остатки только на складе 1619436
            TARGET_WAREHOUSE_ID = 1619436
            
            if TARGET_WAREHOUSE_ID not in all_stocks_data:
                all_stocks_data[TARGET_WAREHOUSE_ID] = []
            
            # Получаем баркод для обновления остатков из файла соответствия (колонка G)
            # Баркод всегда берем из файла "Баркоды.xlsx", так как артикул уже проверен
            barcode_for_stock = None
            if manufacturer_art in manufacturer_art_to_barcode:
                barcode_for_stock = manufacturer_art_to_barcode[manufacturer_art]
            elif manufacturer_art_clean in manufacturer_art_to_barcode:
                barcode_for_stock = manufacturer_art_to_barcode[manufacturer_art_clean]
            else:
                # Пробуем нормализованный вариант
                for art_key, barcode_val in manufacturer_art_to_barcode.items():
                    art_key_normalized = str(art_key).strip().replace(' ', '').upper().replace('-', '').replace('/', '').replace('_', '')
                    if art_key_normalized == manufacturer_art_normalized:
                        barcode_for_stock = barcode_val
                        break
            
            if barcode_for_stock:
                # Используем только sku - API сам найдет chrtId по sku при обновлении остатков
                # Это соответствует логике из update_prices_stocks_wb.py
                all_stocks_data[TARGET_WAREHOUSE_ID].append({
                    "sku": barcode_for_stock,
                    "amount": product['amount']
                })
        
        if matched_count > 0:
            print(f"  {brand}: обработано {matched_count} товаров")
    
    # Автоматическая загрузка Excel шаблона (выполняется независимо от наличия данных)
    if Config.AUTO_DOWNLOAD_EXCEL:
        print("\n[INFO] Автоматическая загрузка Excel шаблона WB...")
        downloaded_file = download_excel_template_automated()
        if downloaded_file:
            print(f"  [OK] Свежий шаблон скачан: {os.path.basename(downloaded_file)}")
    
    if not all_stocks_data and not all_prices_data:
        print("\n[WARN] Не найдено данных для обновления")
        return
    
    # Автоматическая корректировка цен через Excel шаблон WB
    # Примечание: API WB не предоставляет рекомендуемые цены, поэтому используем Excel файл
    if all_prices_data and Config.AUTO_ADJUST_PRICES:
        print("\n[INFO] Автоматическая корректировка цен через Excel шаблон WB...")
        
        # Пытаемся автоматически скачать свежий Excel шаблон (если включено)
        downloaded_file = download_excel_template_automated()
        if downloaded_file:
            print(f"  [OK] Свежий шаблон скачан: {os.path.basename(downloaded_file)}")
        
        # Сначала корректируем Excel файл (J = N - 1)
        auto_adjust_wb_template_prices()
        
        # Затем читаем рекомендуемые цены из Excel и применяем к данным для обновления
        recommended_prices = read_prices_from_excel_template()
        
        if recommended_prices:
            print(f"  [INFO] Применяю корректировку цен на основе Excel шаблона...")
            adjusted_count = 0
            
            for item in all_prices_data:
                nmid = item["nmID"]
                base_price = item.get("base_price", item["price"])
                
                if nmid in recommended_prices:
                    # Устанавливаем цену на 1 меньше рекомендуемой
                    recommended_price = recommended_prices[nmid]
                    adjusted_price = recommended_price - 1
                    # Защита от слишком низких цен
                    min_price = int(base_price * 0.9)
                    adjusted_price = max(adjusted_price, min_price)
                    
                    item["price"] = adjusted_price
                    if adjusted_price != base_price:
                        adjusted_count += 1
            
            if adjusted_count > 0:
                print(f"  [OK] Скорректировано цен: {adjusted_count} (на 1 меньше рекомендуемой из Excel)")
        else:
            print("  [WARN] Рекомендуемые цены не прочитаны из Excel, используем базовые цены")
    elif all_prices_data and not Config.AUTO_ADJUST_PRICES:
        print("\n[INFO] Автоматическая корректировка цен отключена (AUTO_ADJUST_PRICES=false)")
    
    # Выводим информацию о том, что будет обновлено
    total_stocks = sum(len(stocks) for stocks in all_stocks_data.values())
    print(f"\nОбновляю: остатков {total_stocks}, цен {len(all_prices_data)}")
    
    # Обновляем остатки только на складе 1619436
    TARGET_WAREHOUSE_ID = 1619436
    
    if TARGET_WAREHOUSE_ID in all_stocks_data:
        stocks_data = all_stocks_data[TARGET_WAREHOUSE_ID]
        warehouse = next((w for w in warehouses if w.get('id') == TARGET_WAREHOUSE_ID), None)
        warehouse_name = warehouse.get('name', 'Неизвестный склад') if warehouse else 'Неизвестный склад'
        
        # Разбиваем на батчи по 100
        batch_size = 100
        total_batches = (len(stocks_data) + batch_size - 1) // batch_size
        for i in range(0, len(stocks_data), batch_size):
            batch = stocks_data[i:i + batch_size]
            batch_num = i//batch_size + 1
            # Показываем прогресс каждые 10 батчей или последний батч
            if batch_num % 10 == 0 or batch_num == total_batches:
                print(f"  Остатки: батч {batch_num}/{total_batches}...")
            if not update_stocks(TARGET_WAREHOUSE_ID, batch):
                # Если ошибка, делаем задержку перед следующим батчем
                if i + batch_size < len(stocks_data):
                    time.sleep(3)
            
            # Добавляем небольшую задержку между батчами для избежания 429 ошибок
            if i + batch_size < len(stocks_data):
                time.sleep(0.5)
    else:
        print(f"  [WARN] Нет данных для обновления остатков на складе {TARGET_WAREHOUSE_ID}")
    
    # Обновляем цены
    if all_prices_data:
        # Разбиваем на батчи по 100
        batch_size = 100
        total_batches = (len(all_prices_data) + batch_size - 1) // batch_size
        for i in range(0, len(all_prices_data), batch_size):
            batch = all_prices_data[i:i + batch_size]
            batch_num = i//batch_size + 1
            # Показываем прогресс каждые 10 батчей или последний батч
            if batch_num % 10 == 0 or batch_num == total_batches:
                print(f"  Цены: батч {batch_num}/{total_batches}...")
            if not update_prices(batch):
                # Если ошибка, делаем задержку перед следующим батчем
                if i + batch_size < len(all_prices_data):
                    time.sleep(3)
            
            # Добавляем небольшую задержку между батчами для избежания 429 ошибок
            if i + batch_size < len(all_prices_data):
                time.sleep(0.5)
    
    print("Обновление завершено!")


if __name__ == "__main__":
    main()

