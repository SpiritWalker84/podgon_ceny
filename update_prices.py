"""
Скрипт для автоматической корректировки цен в Excel файлах WB.
Устанавливает цену в колонке J на 1 меньше, чем в колонке N.
Это нужно для избежания "непривлекательных цен" на Wildberries.

Использование:
1. Как скрипт: python update_prices.py [путь_к_файлу.xlsx]
2. Как модуль: from update_prices import adjust_prices
"""

import openpyxl
from openpyxl import load_workbook
import sys
import os
import glob
from pathlib import Path


def adjust_prices(file_path, column_n=14, column_j=10, verbose=False):
    """
    Корректирует цены в Excel файле: устанавливает колонку J = N - 1
    
    Args:
        file_path: Путь к Excel файлу
        column_n: Номер колонки N (по умолчанию 14)
        column_j: Номер колонки J (по умолчанию 10)
        verbose: Показывать детальный вывод
    
    Returns:
        Количество измененных строк
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл не найден: {file_path}")
    
    # Открываем файл
    wb = load_workbook(file_path)
    ws = wb.active
    
    max_row = ws.max_row
    changes_count = 0
    
    for row in range(1, max_row + 1):
        # Читаем значение из колонки N
        cell_n = ws.cell(row=row, column=column_n)
        cell_j = ws.cell(row=row, column=column_j)
        
        # Проверяем, есть ли значение в колонке N
        if cell_n.value is not None:
            try:
                # Пытаемся преобразовать в число
                value_n = float(cell_n.value)
                # Вычисляем новое значение для J (на 1 меньше)
                new_value_j = value_n - 1
                
                # Записываем новое значение в колонку J
                cell_j.value = new_value_j
                changes_count += 1
                
                if verbose:
                    print(f"Строка {row}: N={value_n} -> J={new_value_j}")
            except (ValueError, TypeError):
                # Если значение не является числом, пропускаем
                if verbose:
                    print(f"Строка {row}: значение в N не является числом ({cell_n.value}), пропускаем")
    
    # Сохраняем файл
    wb.save(file_path)
    return changes_count


def find_wb_template_files(directory="."):
    """
    Находит файлы шаблонов WB по паттерну имени
    
    Args:
        directory: Директория для поиска (по умолчанию текущая)
    
    Returns:
        Список найденных файлов
    """
    patterns = [
        "Шаблон обновления цен и скидок*.xlsx",
        "*обновления цен*.xlsx",
        "*WB*.xlsx",
    ]
    
    found_files = []
    for pattern in patterns:
        found_files.extend(glob.glob(os.path.join(directory, pattern)))
    
    # Удаляем дубликаты и сортируем по дате модификации (новые первыми)
    found_files = list(set(found_files))
    found_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    
    return found_files


def main():
    """Основная функция для запуска скрипта из командной строки"""
    if len(sys.argv) > 1:
        # Если передан путь к файлу как аргумент
        file_path = sys.argv[1]
        if not os.path.exists(file_path):
            print(f"Ошибка: Файл не найден: {file_path}")
            sys.exit(1)
    else:
        # Автоматически ищем файл шаблона
        found_files = find_wb_template_files()
        if not found_files:
            print("Ошибка: Не найден файл шаблона WB.")
            print("Использование: python update_prices.py [путь_к_файлу.xlsx]")
            sys.exit(1)
        
        # Берем самый новый файл
        file_path = found_files[0]
        print(f"Найден файл: {file_path}")
    
    try:
        print(f"Обработка файла: {file_path}")
        changes_count = adjust_prices(file_path, verbose=True)
        print(f"\nГотово! Изменено {changes_count} строк.")
    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

